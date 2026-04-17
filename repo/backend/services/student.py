"""Student records and search."""
from __future__ import annotations
import csv
import uuid
from dataclasses import asdict
from datetime import date, datetime
from pathlib import Path

from .. import audit, crypto, db, events
from ..models import (ImportPreview, Student, StudentDTO, StudentSummary,
                      ChangeLogEntry)
from ..permissions import Session, requires
from .auth import BizError


_PREVIEW_CACHE: dict[str, ImportPreview] = {}


def _redact(dto: StudentDTO) -> dict:
    """Return an audit-safe dict — PII fields replaced with masked tokens."""
    d = asdict(dto)
    if d.get("email"):
        d["email"] = "***@***"
    if d.get("phone"):
        d["phone"] = "***-****"
    if d.get("ssn_last4"):
        d["ssn_last4"] = "***"
    return d


class StudentService:

    REQUIRED_COLUMNS = ["student_id", "full_name", "college", "class_year",
                        "email", "phone", "housing_status"]

    _SEARCH_PERMS = ("student.write", "student.import", "student.pii.read",
                     "housing.write", "system.admin")

    # ---- read -------------------------------------------------------------

    def get(self, session: Session, student_id: int) -> Student:
        if not session.has_any(self._SEARCH_PERMS):
            from ..permissions import PermissionDenied
            raise PermissionDenied("student.read")
        conn = db.get_connection()
        row = conn.execute(
            "SELECT * FROM students WHERE id = ?", (student_id,)).fetchone()
        if not row:
            raise BizError("STUDENT_NOT_FOUND", "Student not found.")
        return self._row_to_student(session, row)

    def search(self, session: Session, *, text: str | None = None,
               college: str | None = None, housing_status: str | None = None,
               page: "Page | None" = None,
               limit: int | None = None,
               offset: int | None = None) -> "Paged[StudentSummary]":
        from ..models import Page, Paged
        if page is None:
            page = Page()
        # Convenience: callers may pass ``limit=`` / ``offset=`` directly
        # without constructing a Page object.
        if limit is not None:
            page = Page(limit=int(limit),
                        offset=int(offset) if offset is not None else page.offset)
        elif offset is not None:
            page = Page(limit=page.limit, offset=int(offset))
        if not session.has_any(self._SEARCH_PERMS):
            from ..permissions import PermissionDenied
            raise PermissionDenied("student.read")
        conn = db.get_connection()
        where, args = [], []
        if text:
            where.append("(full_name LIKE ? OR student_id_ext LIKE ?)")
            like = f"%{text}%"
            args += [like, like]
        if college:
            where.append("college = ?")
            args.append(college)
        if housing_status:
            where.append("housing_status = ?")
            args.append(housing_status)
        base_sql = "FROM students"
        if where:
            base_sql += " WHERE " + " AND ".join(where)
        total = conn.execute("SELECT COUNT(*) " + base_sql, args).fetchone()[0]
        sql = "SELECT id, student_id_ext, full_name, college, class_year, housing_status " + base_sql
        sql += " ORDER BY full_name LIMIT ? OFFSET ?"
        rows = conn.execute(sql, args + [page.limit, page.offset]).fetchall()
        items = [StudentSummary(
            id=r["id"], student_id=r["student_id_ext"], full_name=r["full_name"],
            college=r["college"], class_year=r["class_year"],
            housing_status=r["housing_status"]) for r in rows]
        return Paged(items=items, total=total)

    # ---- write ------------------------------------------------------------

    @requires("student.write")
    def create(self, session: Session, dto: StudentDTO) -> Student:
        self._validate(dto)
        conn = db.get_connection()
        existing = conn.execute(
            "SELECT 1 FROM students WHERE student_id_ext = ?",
            (dto.student_id,)).fetchone()
        if existing:
            raise BizError("STUDENT_DUPLICATE_ID",
                           f"Student ID {dto.student_id} already exists.")
        with db.transaction() as conn:
            cur = conn.execute(
                """INSERT INTO students(student_id_ext, full_name, college,
                       class_year, email_enc, phone_enc, ssn_last4_enc,
                       housing_status)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
                (dto.student_id, dto.full_name, dto.college, dto.class_year,
                 crypto.encrypt_field(dto.email),
                 crypto.encrypt_field(dto.phone),
                 crypto.encrypt_field(dto.ssn_last4),
                 dto.housing_status))
            new_id = cur.lastrowid
            conn.execute(
                "INSERT INTO students_fts(rowid, student_id_ext, full_name, college) "
                "VALUES (?, ?, ?, ?)",
                (new_id, dto.student_id, dto.full_name, dto.college or ""))
        audit.record(session.user_id, "student", new_id, "create", _redact(dto))
        events.bus.publish(events.STUDENT_CREATED,
                           {"student_id": new_id, "name": dto.full_name})
        return self.get(session, new_id)

    @requires("student.write")
    def update(self, session: Session, student_id: int, dto: StudentDTO) -> Student:
        self._validate(dto)
        with db.transaction() as conn:
            # Read old FTS column values before mutating — contentless FTS5
            # requires the previous values to remove the old entry.
            old = conn.execute(
                "SELECT student_id_ext, full_name, college FROM students WHERE id=?",
                (student_id,)).fetchone()
            conn.execute(
                """UPDATE students SET student_id_ext=?, full_name=?, college=?,
                       class_year=?, email_enc=?, phone_enc=?, ssn_last4_enc=?,
                       housing_status=?, updated_at=datetime('now')
                   WHERE id = ?""",
                (dto.student_id, dto.full_name, dto.college, dto.class_year,
                 crypto.encrypt_field(dto.email),
                 crypto.encrypt_field(dto.phone),
                 crypto.encrypt_field(dto.ssn_last4),
                 dto.housing_status, student_id))
            # Contentless FTS5 (content='') does not support plain DELETE;
            # use the special 'delete' command with old column values.
            if old:
                conn.execute(
                    "INSERT INTO students_fts(students_fts, rowid, student_id_ext, "
                    "full_name, college) VALUES('delete', ?, ?, ?, ?)",
                    (student_id, old["student_id_ext"], old["full_name"],
                     old["college"] or ""))
            conn.execute(
                "INSERT INTO students_fts(rowid, student_id_ext, full_name, college) "
                "VALUES (?, ?, ?, ?)",
                (student_id, dto.student_id, dto.full_name, dto.college or ""))
        audit.record(session.user_id, "student", student_id, "update", _redact(dto))
        events.bus.publish(events.STUDENT_UPDATED,
                           {"student_id": student_id, "name": dto.full_name})
        return self.get(session, student_id)

    # ---- import / export --------------------------------------------------

    @requires("student.import")
    def import_file(self, session: Session, path: str | Path,
                    duplicate_strategy: str = "error") -> ImportPreview:
        """Dry-run preview for either CSV (.csv) or Excel (.xlsx)."""
        if duplicate_strategy not in ("skip", "update", "error"):
            raise BizError("BAD_OPTION", "duplicate_strategy invalid")
        ext = Path(path).suffix.lower()
        if ext == ".xlsx":
            rows_iter, cols = self._iter_xlsx(path)
        elif ext == ".csv":
            rows_iter, cols = self._iter_csv(path)
        else:
            raise BizError("BAD_FORMAT",
                           f"Unsupported file type: {ext} (expected .csv or .xlsx)")

        missing = [c for c in self.REQUIRED_COLUMNS if c not in cols]
        if missing:
            raise BizError("BAD_HEADER",
                           f"Missing required columns: {', '.join(missing)}")

        accepted, rejected = [], []
        seen: set[str] = set()
        conn = db.get_connection()
        existing_ids = {r["student_id_ext"] for r in
                        conn.execute("SELECT student_id_ext FROM students")}
        for i, row in rows_iter:
            err = self._validate_row(row)
            if err:
                rejected.append({"row": i, "data": row, "error": err})
                continue
            sid = (row.get("student_id") or "").strip()
            if sid in seen:
                rejected.append({"row": i, "data": row,
                                 "error": "duplicate ID within file"})
                continue
            seen.add(sid)
            if sid in existing_ids:
                if duplicate_strategy == "error":
                    rejected.append({"row": i, "data": row,
                                     "error": "duplicate of existing record"})
                    continue
                row["_action"] = "update" if duplicate_strategy == "update" else "skip"
            else:
                row["_action"] = "create"
            accepted.append({"row": i, "data": row})
        preview = ImportPreview(
            preview_id=str(uuid.uuid4()),
            accepted=accepted,
            rejected=rejected,
            columns=self.REQUIRED_COLUMNS,
            duplicate_strategy=duplicate_strategy,
        )
        _PREVIEW_CACHE[preview.preview_id] = preview
        return preview

    @requires("student.import")
    def commit_import(self, session: Session, preview_id: str) -> dict:
        preview = _PREVIEW_CACHE.pop(preview_id, None)
        if not preview:
            raise BizError("PREVIEW_EXPIRED", "Import preview not found or expired.")
        created, updated, skipped = 0, 0, 0
        for entry in preview.accepted:
            row = entry["data"]
            action = row.get("_action")
            dto = StudentDTO(
                student_id=row["student_id"].strip(),
                full_name=row["full_name"].strip(),
                college=row.get("college") or None,
                class_year=int(row["class_year"]) if row.get("class_year") else None,
                email=row.get("email") or None,
                phone=row.get("phone") or None,
                ssn_last4=row.get("ssn_last4") or None,
                housing_status=row.get("housing_status") or "pending",
            )
            try:
                if action == "create":
                    self.create(session, dto)
                    created += 1
                elif action == "update":
                    conn = db.get_connection()
                    rec = conn.execute(
                        "SELECT id FROM students WHERE student_id_ext=?",
                        (dto.student_id,)).fetchone()
                    if rec:
                        self.update(session, rec["id"], dto)
                        updated += 1
                else:
                    skipped += 1
            except Exception:
                skipped += 1
        return {"created": created, "updated": updated, "skipped": skipped,
                "rejected": len(preview.rejected)}

    # Back-compat alias used by some callers and tests.
    def import_csv(self, session: Session, path: str | Path,
                   duplicate_strategy: str = "error") -> ImportPreview:
        return self.import_file(session, path, duplicate_strategy)

    @requires("student.import")
    def export_file(self, session: Session, path: str | Path,
                    include_pii: bool = False) -> int:
        """Export to CSV (.csv) or Excel (.xlsx) — chosen by file extension."""
        rows = list(self.search(session, limit=10_000))
        full_rows: list[list] = []
        for r in rows:
            full = self.get(session, r.id)
            full_rows.append([
                full.student_id, full.full_name, full.college or "",
                full.class_year or "", full.email or "", full.phone or "",
                full.housing_status,
            ])
        ext = Path(path).suffix.lower()
        if ext == ".xlsx":
            try:
                from openpyxl import Workbook
            except ImportError as e:
                raise BizError("XLSX_UNAVAILABLE",
                               "openpyxl not installed") from e
            wb = Workbook()
            ws = wb.active
            ws.title = "Students"
            ws.append(self.REQUIRED_COLUMNS)
            for r in full_rows:
                ws.append(r)
            ws.freeze_panes = "A2"
            wb.save(str(path))
        elif ext == ".csv":
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(self.REQUIRED_COLUMNS)
                for r in full_rows:
                    writer.writerow(r)
        else:
            raise BizError("BAD_FORMAT",
                           f"Unsupported export type: {ext} (.csv or .xlsx)")
        return len(rows)

    # Back-compat alias.
    def export_csv(self, session: Session, path: str | Path,
                   include_pii: bool = False) -> int:
        return self.export_file(session, path, include_pii=include_pii)

    # ---- format readers ------------------------------------------------

    def _iter_csv(self, path: str | Path):
        with open(path, newline="", encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            cols = reader.fieldnames or []
            data = [(i, dict(row)) for i, row in enumerate(reader, start=2)]
        return data, cols

    def _iter_xlsx(self, path: str | Path):
        try:
            from openpyxl import load_workbook
        except ImportError as e:
            raise BizError("XLSX_UNAVAILABLE",
                           "openpyxl is required to read .xlsx files") from e
        wb = load_workbook(filename=str(path), read_only=True, data_only=True)
        ws = wb.active
        rows = ws.iter_rows(values_only=True)
        try:
            header = [str(c).strip() if c is not None else ""
                      for c in next(rows)]
        except StopIteration:
            raise BizError("BAD_FORMAT", "Empty workbook")
        data: list = []
        for i, raw in enumerate(rows, start=2):
            row = {header[j]: ("" if v is None else str(v).strip())
                   for j, v in enumerate(raw) if j < len(header)}
            if any(row.values()):
                data.append((i, row))
        return data, header

    # ---- history ----------------------------------------------------------

    def history(self, session: Session, student_id: int) -> list[ChangeLogEntry]:
        if not session.has_any(self._SEARCH_PERMS):
            from ..permissions import PermissionDenied
            raise PermissionDenied("student.read")
        conn = db.get_connection()
        rows = conn.execute(
            "SELECT id, ts, actor_id, action, payload_json FROM audit_log "
            "WHERE entity_type='student' AND entity_id=? ORDER BY id DESC",
            (str(student_id),)).fetchall()
        import json
        return [ChangeLogEntry(
            id=r["id"], ts=r["ts"], actor_id=r["actor_id"], action=r["action"],
            payload=json.loads(r["payload_json"])) for r in rows]

    # ---- helpers ----------------------------------------------------------

    def _row_to_student(self, session: Session, row) -> Student:
        unlocked = session.mask_unlocked() and session.has("student.pii.read")
        email = crypto.decrypt_field(row["email_enc"])
        phone = crypto.decrypt_field(row["phone_enc"])
        ssn = crypto.decrypt_field(row["ssn_last4_enc"])
        if not unlocked:
            email = crypto.mask_email(email)
            phone = crypto.mask_phone(phone)
            ssn = crypto.mask_ssn_last4(ssn)
        return Student(
            id=row["id"], student_id=row["student_id_ext"],
            full_name=row["full_name"], college=row["college"],
            class_year=row["class_year"], email=email, phone=phone,
            ssn_last4=ssn, housing_status=row["housing_status"],
            created_at=row["created_at"], updated_at=row["updated_at"],
        )

    def _validate(self, dto: StudentDTO) -> None:
        if not dto.student_id or not dto.full_name:
            raise BizError("MISSING_FIELD", "student_id and full_name are required.")
        if dto.housing_status not in ("on_campus", "off_campus", "pending"):
            raise BizError("BAD_STATUS", "housing_status invalid.")

    def _validate_row(self, row: dict) -> str | None:
        if not row.get("student_id"):
            return "student_id required"
        if not row.get("full_name"):
            return "full_name required"
        cy = row.get("class_year")
        if cy:
            try:
                int(cy)
            except ValueError:
                return "class_year must be integer"
        # Date-format validator (MM/DD/YYYY) for any *_date column.
        for k, v in row.items():
            if k.endswith("_date") and v:
                try:
                    datetime.strptime(v, "%m/%d/%Y")
                except ValueError:
                    return f"{k} must be MM/DD/YYYY"
        return None
